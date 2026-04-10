#!/usr/bin/env python3
"""
youtube_research.py — YouTube リサーチをトークン消費ゼロで実行

YouTube Data API v3 を直接呼び出し、結果をExcel(.xlsx)に出力する。
Next.jsサーバー不要。AIへのトークン消費なし。

出力命名規則:
  Excel: YTR_{英字KW}_{YYYYMMDD}_{連番}.xlsx
  コメントJSON: YTR_{英字KW}_{YYYYMMDD}_{連番}_comments.json
  例: YTR_AI_Document_creation_20260402_1.xlsx / _2.xlsx / _3.xlsx

Usage:
  python3 youtube_research.py --keywords "AI活用,Claude使い方" --ep EP-05
  python3 youtube_research.py --ep EP-05          # 企画骨子から自動抽出
  python3 youtube_research.py --keywords "AI活用" # 単一KW（対話で拡張可）

Options:
  --keywords    カンマ区切りキーワード
  --ep          EP番号（例: EP-05）。GutxxxBrainの企画骨子からKWを自動抽出
  --period      検索期間（月数、デフォルト6）
  --type        動画タイプ: any / normal / shorts（デフォルト: normal）
  --no-comments コメント取得をスキップ
  --output-dir  Excel・JSON出力先ディレクトリ（デフォルト: ./output）
"""

import re
import sys
import json
import argparse
import unicodedata
import requests
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# 設定
# ============================================================

SCRIPT_DIR = Path(__file__).parent
ENV_PATH = SCRIPT_DIR / ".env"
GUTXXXBRAIN_PATH = Path("/Users/daisukemiyauchi/Documents/GitHub/GutxxxBrain")

YT_BASE = "https://www.googleapis.com/youtube/v3"

# ============================================================
# .env 読み込み / APIキー収集
# ============================================================

def load_env(path: Path) -> dict:
    env = {}
    if not path.exists():
        return env
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            key, _, val = line.partition("=")
            env[key.strip()] = val.strip()
    return env

def collect_api_keys(env: dict) -> list:
    keys = []
    if env.get("YOUTUBE_API_KEY"):
        keys.append(env["YOUTUBE_API_KEY"])
    for i in range(1, 11):
        key = env.get(f"YOUTUBE_API_KEY_{i}", "")
        if key and "のキー" not in key:
            keys.append(key)
    return keys

# ============================================================
# YouTube API クライアント
# ============================================================

class YouTubeClient:
    def __init__(self, api_keys: list):
        if not api_keys:
            raise RuntimeError("YouTube API キーが設定されていません。.env を確認してください。")
        self.api_keys = api_keys
        self.key_index = 0

    def _get(self, endpoint: str, params: dict) -> dict:
        tried = set()
        last_error = None
        while len(tried) < len(self.api_keys):
            tried.add(self.key_index)
            params["key"] = self.api_keys[self.key_index]
            try:
                r = requests.get(f"{YT_BASE}/{endpoint}", params=params, timeout=30)
                if r.status_code in (403, 429):
                    print(f"  ⚠ APIキー #{self.key_index + 1} クォータ超過 → 次のキーへ切り替え")
                    last_error = RuntimeError(f"HTTP {r.status_code}")
                    self.key_index = (self.key_index + 1) % len(self.api_keys)
                    continue
                r.raise_for_status()
                return r.json()
            except requests.RequestException as e:
                raise RuntimeError(f"API リクエスト失敗: {e}")
        raise RuntimeError(f"全APIキーがクォータ超過: {last_error}")

    def search(self, query: str, start_date: str, end_date: str) -> list:
        """検索（最大3ページ = 150件）"""
        all_items = []
        next_page = None
        for _ in range(3):
            params = {
                "part": "snippet",
                "q": query,
                "type": "video",
                "regionCode": "JP",
                "relevanceLanguage": "ja",
                "order": "relevance",
                "maxResults": 50,
                "publishedAfter": f"{start_date}T00:00:00Z",
                "publishedBefore": f"{end_date}T23:59:59Z",
            }
            if next_page:
                params["pageToken"] = next_page
            data = self._get("search", params)
            all_items.extend(data.get("items", []))
            next_page = data.get("nextPageToken")
            if not next_page:
                break
        # 重複除去
        seen = set()
        unique = []
        for item in all_items:
            vid = item["id"]["videoId"]
            if vid not in seen:
                seen.add(vid)
                unique.append(item)
        return unique

    def get_video_details(self, video_ids: list) -> dict:
        result = {}
        for i in range(0, len(video_ids), 50):
            batch = video_ids[i:i + 50]
            data = self._get("videos", {
                "part": "statistics,contentDetails",
                "id": ",".join(batch),
            })
            for item in data.get("items", []):
                result[item["id"]] = item
        return result

    def get_channel_details(self, channel_ids: list) -> dict:
        result = {}
        unique_ids = list(dict.fromkeys(channel_ids))
        for i in range(0, len(unique_ids), 50):
            batch = unique_ids[i:i + 50]
            data = self._get("channels", {
                "part": "snippet,statistics",
                "id": ",".join(batch),
            })
            for item in data.get("items", []):
                result[item["id"]] = item
        return result

    def get_comments(self, video_id: str) -> list:
        threads = []
        next_page = None
        for _ in range(20):
            params = {
                "part": "snippet,replies",
                "videoId": video_id,
                "maxResults": 100,
                "textFormat": "plainText",
                "order": "relevance",
            }
            if next_page:
                params["pageToken"] = next_page
            try:
                data = self._get("commentThreads", params)
            except Exception as e:
                print(f"  コメント取得エラー ({video_id}): {e}")
                break
            for item in data.get("items", []):
                top = item["snippet"]["topLevelComment"]["snippet"]
                thread = {
                    "id": item["id"],
                    "author": top["authorDisplayName"],
                    "text": top["textDisplay"],
                    "likes": top["likeCount"],
                    "publishedAt": top["publishedAt"],
                    "replies": [],
                }
                for reply in item.get("replies", {}).get("comments", []):
                    rs = reply["snippet"]
                    thread["replies"].append({
                        "author": rs["authorDisplayName"],
                        "text": rs["textDisplay"],
                        "likes": rs["likeCount"],
                        "publishedAt": rs["publishedAt"],
                    })
                threads.append(thread)
            next_page = data.get("nextPageToken")
            if not next_page:
                break
        return threads

# ============================================================
# フィルター / 判定ロジック（filter.ts + youtube.ts から移植）
# ============================================================

JP_REGEX = re.compile(r'[\u3040-\u309F\u30A0-\u30FF]')
SHORTS_TAG = re.compile(r'#shorts?\b', re.IGNORECASE)

def parse_duration(iso: str) -> int:
    m = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', iso or "")
    if not m:
        return 0
    return int(m.group(1) or 0) * 3600 + int(m.group(2) or 0) * 60 + int(m.group(3) or 0)

def format_duration(iso: str) -> str:
    s = parse_duration(iso)
    if s == 0:
        return ""
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    return f"{h}:{m:02d}:{sec:02d}" if h > 0 else f"{m}:{sec:02d}"

def is_short(duration: str, title: str, desc: str) -> bool:
    s = parse_duration(duration)
    if s == 0:
        return False
    if s > 180:
        return False
    if s <= 60:
        return True
    return bool(SHORTS_TAG.search(title + " " + desc))

def is_domestic(country: str, title: str, desc: str) -> bool:
    if country == "JP":
        return True
    if country and country != "JP":
        return False
    return bool(JP_REGEX.search(title + " " + desc))

# ============================================================
# 動画データ構築
# ============================================================

def build_items(search_items: list, video_details: dict, channel_details: dict, video_type: str) -> list:
    now = datetime.now()
    items = []
    for s in search_items:
        vid_id = s["id"]["videoId"]
        ch_id = s["snippet"]["channelId"]
        vd = video_details.get(vid_id, {})
        cd = channel_details.get(ch_id, {})
        stats = vd.get("statistics", {})
        duration = vd.get("contentDetails", {}).get("duration", "")
        title = s["snippet"]["title"]
        desc = s["snippet"]["description"]
        country = cd.get("snippet", {}).get("country", "")

        if not is_domestic(country, title, desc):
            continue

        short = is_short(duration, title, desc)
        if video_type == "normal" and short:
            continue
        if video_type == "shorts" and not short:
            continue

        view = int(stats.get("viewCount", 0))
        like = int(stats.get("likeCount", 0))
        comment = int(stats.get("commentCount", 0))
        sub = int(cd.get("statistics", {}).get("subscriberCount", 0))

        pub = datetime.fromisoformat(s["snippet"]["publishedAt"].replace("Z", "+00:00"))
        days = max(1, (now - pub.replace(tzinfo=None)).days)

        thumb = (
            s["snippet"]["thumbnails"].get("medium")
            or s["snippet"]["thumbnails"].get("high")
            or {}
        ).get("url", "")

        items.append({
            "videoId": vid_id,
            "title": title,
            "channelTitle": s["snippet"]["channelTitle"],
            "publishedAt": pub.strftime("%Y/%m/%d"),
            "subscriberCount": sub,
            "viewCount": view,
            "likeCount": like,
            "commentCount": comment,
            "videoUrl": f"https://www.youtube.com/watch?v={vid_id}",
            "thumbnailUrl": thumb,
            "duration": duration,
            "durationFmt": format_duration(duration),
            "likeRate": like / view if view > 0 else 0,
            "commentRate": comment / view if view > 0 else 0,
            "viewSubRatio": view / sub if sub > 0 else 0,
            "dailyAvg": view / days,
        })

    items.sort(key=lambda x: x["viewCount"], reverse=True)
    return items

# ============================================================
# Excel 出力
# ============================================================

RESEARCH_COLS = [
    ("調査日時", 22), ("検索KW", 16), ("タイトル", 45), ("公開日", 12),
    ("登録者数", 12), ("再生数", 12), ("いいね数", 10), ("コメント数", 12),
    ("高評価率", 10), ("コメント率", 10), ("V/S比", 9), ("日次平均再生", 13),
    ("動画の尺", 10), ("チャンネル名", 22), ("動画URL", 40), ("サムネイルURL", 45),
]
COMMENT_COLS = [
    ("動画ID", 16), ("動画タイトル", 32), ("タイプ", 9), ("投稿者", 18),
    ("投稿日", 12), ("いいね数", 10), ("コメント本文", 55), ("動画URL", 40),
]

HEADER_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
DATA_ALIGN = Alignment(vertical="center", wrap_text=True)


def _style_header(ws, cols):
    for i, (name, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def write_research_sheet(ws, items: list, keyword: str):
    _style_header(ws, RESEARCH_COLS)
    now_str = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    for r, item in enumerate(items, 2):
        row_data = [
            now_str,
            keyword,
            item["title"],
            item["publishedAt"],
            item["subscriberCount"],
            item["viewCount"],
            item["likeCount"],
            item["commentCount"],
            f'{item["likeRate"] * 100:.2f}%',
            f'{item["commentRate"] * 100:.3f}%',
            f'{item["viewSubRatio"]:.2f}',
            round(item["dailyAvg"]),
            item["durationFmt"],
            item["channelTitle"],
            item["videoUrl"],
            item["thumbnailUrl"] or "",
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = DATA_ALIGN


def write_comments_sheet(ws, comments_data: list):
    _style_header(ws, COMMENT_COLS)
    r = 2
    for entry in comments_data:
        row_data = [
            entry["videoId"], entry["videoTitle"], "Thread",
            entry["author"], entry["publishedAt"][:10],
            entry["likes"], entry["text"], entry["videoUrl"],
        ]
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val).alignment = DATA_ALIGN
        r += 1
        for reply in entry.get("replies", []):
            row_data = [
                entry["videoId"], entry["videoTitle"], "Reply",
                reply["author"], reply["publishedAt"][:10],
                reply["likes"], reply["text"], entry["videoUrl"],
            ]
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=val).alignment = DATA_ALIGN
            r += 1


# ============================================================
# KW → 英字ファイル名変換
# ============================================================

# 日本語KWから英字ファイル名を生成する簡易マッピング
# よく使うKWを事前定義。未登録のKWはスペース→アンダースコア変換
KW_EN_MAP = {
    "AI": "AI",
    "ai": "AI",
    "資料作成": "Document_creation",
    "プレゼン": "Presentation",
    "自動化": "Automation",
    "効率化": "Efficiency",
    "使い方": "How_to_use",
    "活用": "Usage",
    "業務": "Business",
    "仕事": "Work",
    "ツール": "Tools",
    "比較": "Comparison",
    "まとめ": "Summary",
    "初心者": "Beginner",
    "入門": "Introduction",
    "最新": "Latest",
    "おすすめ": "Recommended",
    "無料": "Free",
    "有料": "Paid",
    "動画編集": "Video_editing",
    "画像生成": "Image_generation",
    "音声": "Audio",
    "文章": "Writing",
    "翻訳": "Translation",
    "プログラミング": "Programming",
    "コーディング": "Coding",
    "デザイン": "Design",
    "マーケティング": "Marketing",
    "副業": "Side_business",
    "生産性": "Productivity",
    "議事録": "Meeting_notes",
    "要約": "Summarize",
    "ChatGPT": "ChatGPT",
    "Claude": "Claude",
    "Gemini": "Gemini",
    "Copilot": "Copilot",
    "NotebookLM": "NotebookLM",
    "Notebook LM": "NotebookLM",
}


def _try_submatch(token: str) -> list:
    """トークン内に辞書のキーが部分一致する場合、分解して変換する。
    例: 'AI活用' → 'AI' + '活用' → ['AI', 'Usage']
    """
    # 長いキーから順にマッチを試みる（貪欲マッチ）
    sorted_keys = sorted(KW_EN_MAP.keys(), key=len, reverse=True)
    remaining = token
    parts = []
    while remaining:
        matched = False
        for key in sorted_keys:
            if remaining.startswith(key):
                parts.append(KW_EN_MAP[key])
                remaining = remaining[len(key):]
                matched = True
                break
        if not matched:
            # 1文字ずつ進める（マッチしない部分を蓄積）
            # 次のマッチ開始位置を探す
            found = False
            for i in range(1, len(remaining) + 1):
                for key in sorted_keys:
                    if remaining[i:].startswith(key):
                        # i文字分はマッチしなかった部分
                        unmatched = remaining[:i]
                        normalized = unicodedata.normalize('NFKD', unmatched)
                        ascii_part = normalized.encode('ascii', 'ignore').decode('ascii')
                        if ascii_part:
                            parts.append(ascii_part)
                        remaining = remaining[i:]
                        found = True
                        break
                if found:
                    break
            if not found:
                # 残り全部マッチしない
                return None
    return parts if parts else None


def kw_to_filename_part(kw: str) -> str:
    """日本語KWを英字ファイル名パーツに変換する。

    1. KW全体がマッピングにあればそのまま使う
    2. なければスペースで分割し、各トークンを:
       a. マッピング完全一致
       b. 部分マッチで分解（例: 'AI活用' → 'AI_Usage'）
       c. ASCII化
       d. フォールバック: ハッシュ短縮
    """
    kw = kw.strip()
    if kw in KW_EN_MAP:
        return KW_EN_MAP[kw]

    tokens = re.split(r'[\s　]+', kw)
    parts = []
    for token in tokens:
        if token in KW_EN_MAP:
            parts.append(KW_EN_MAP[token])
        elif re.match(r'^[A-Za-z0-9._-]+$', token):
            parts.append(token)
        else:
            # 部分マッチで分解を試みる
            sub = _try_submatch(token)
            if sub:
                parts.extend(sub)
            else:
                normalized = unicodedata.normalize('NFKD', token)
                ascii_part = normalized.encode('ascii', 'ignore').decode('ascii')
                if ascii_part:
                    parts.append(ascii_part)
                else:
                    parts.append(f"kw{abs(hash(token)) % 10000:04d}")

    result = "_".join(parts)
    result = re.sub(r'[^A-Za-z0-9_]', '_', result)
    result = re.sub(r'_+', '_', result).strip('_')
    return result or "research"


def generate_spreadsheet_name(kw: str, index: int, date_str: str = None) -> str:
    """YTR_{英字KW}_{YYYYMMDD}_{連番} 形式のスプレッドシート名を生成"""
    if date_str is None:
        date_str = datetime.now().strftime("%Y%m%d")
    en_part = kw_to_filename_part(kw)
    return f"YTR_{en_part}_{date_str}_{index}"


# ============================================================
# 企画骨子からキーワード抽出
# ============================================================

def extract_keywords_from_plan(plan_path: Path) -> list:
    text = plan_path.read_text()
    keywords = []
    in_kw_section = False
    for line in text.splitlines():
        if re.search(r'STEP\s*2|キーワード', line, re.IGNORECASE):
            in_kw_section = True
        if in_kw_section and line.startswith("|"):
            cells = [c.strip() for c in line.split("|") if c.strip()]
            if not cells or re.match(r'^[-:]+$', cells[0]):
                continue
            if cells[0] in ("キーワード", "KW", "keyword"):
                continue
            keywords.append(cells[0])
    return keywords


def find_plan_file(ep: str) -> Path:
    plan_dir = GUTXXXBRAIN_PATH / "Youtube_DAisuke_AI" / "01_Planning" / ep
    if plan_dir.exists():
        for f in sorted(plan_dir.glob("*.md")):
            return f
    return None

# ============================================================
# 関連KW自動候補生成（検索結果タイトルから頻出語抽出）
# ============================================================

# 除外ストップワード（日本語の一般的な助詞・接続詞・汎用語）
STOP_WORDS = {
    "の", "を", "に", "は", "が", "で", "と", "も", "な", "た", "し", "て",
    "する", "した", "です", "ます", "から", "まで", "より", "など", "これ",
    "それ", "あれ", "この", "その", "ある", "ない", "いる", "れる", "られる",
    "こと", "もの", "ため", "よう", "方法", "方", "的", "化", "版", "編",
    "使い方", "紹介", "解説", "徹底", "完全", "最新", "おすすめ", "比較",
    "ランキング", "まとめ", "やり方", "できる", "簡単", "無料", "有料",
    "分", "秒", "時間", "選", "つ", "個", "本", "回", "年", "月", "日",
    "人", "万", "円", "件", "一", "二", "三", "四", "五", "六", "七", "八",
    "九", "十", "百", "千",
}


def extract_trending_keywords(search_items: list, original_kw: str) -> list:
    """検索結果のタイトル・説明文から頻出するトレンドKWを抽出する。

    方法:
    1. 全タイトル＋説明文を収集
    2. 英字トークン（製品名・サービス名）と日本語N-gramを抽出
    3. 元のKWに含まれる語を除外
    4. 出現頻度上位をソートして返す
    """
    texts = []
    for item in search_items:
        texts.append(item["snippet"]["title"])
        texts.append(item["snippet"].get("description", ""))

    combined = " ".join(texts)

    # 英字トークン抽出（製品名・サービス名を拾う: 2文字以上）
    en_tokens = re.findall(r'[A-Za-z][A-Za-z0-9.]+(?:\s[A-Z][A-Za-z0-9.]+)*', combined)
    # 日本語カタカナ語抽出（3文字以上のカタカナ連続）
    katakana_tokens = re.findall(r'[\u30A0-\u30FF]{3,}', combined)
    # 日本語漢字+ひらがな複合語（2〜6文字）
    kanji_tokens = re.findall(r'[\u4E00-\u9FFF\u3040-\u309F]{2,6}', combined)

    # 元KWのトークンを除外用に分解
    original_parts = set(re.split(r'[\s　]+', original_kw.lower()))
    original_parts.add(original_kw.lower())

    counter = Counter()
    for token in en_tokens:
        t = token.strip()
        if len(t) >= 2 and t.lower() not in original_parts:
            # 正規化: 先頭大文字統一
            counter[t] += 1

    for token in katakana_tokens:
        if token not in STOP_WORDS and token not in original_parts:
            counter[token] += 1

    for token in kanji_tokens:
        if token not in STOP_WORDS and token not in original_parts:
            counter[token] += 1

    # 出現2回以上、頻度順に上位20件を候補として返す
    candidates = [
        (word, count) for word, count in counter.most_common(40)
        if count >= 2
    ][:20]

    return candidates


def suggest_related_keywords(yt: "YouTubeClient", main_kw: str, start_str: str, end_str: str) -> list:
    """メインKWで検索し、結果からトレンドKW候補を生成・提示する"""
    print(f"\n「{main_kw}」で予備検索 → トレンドKW候補を抽出中...")
    search_items = yt.search(main_kw, start_str, end_str)
    if not search_items:
        print("  検索結果0件。手動入力に切り替えます。")
        return []

    candidates = extract_trending_keywords(search_items, main_kw)
    if not candidates:
        print("  トレンドKW候補が見つかりませんでした。")
        return []

    print(f"\n■ トレンドKW候補（検索結果 {len(search_items)}件のタイトル/説明文から抽出）:")
    print("-" * 50)
    for i, (word, count) in enumerate(candidates, 1):
        print(f"  {i:2d}. {word} (出現{count}回)")
    print("-" * 50)

    print("\n組み合わせKWを生成します。")
    print("候補番号をカンマ区切りで選択（例: 1,3,5）")
    print("または独自KWを直接入力（例: NotebookLM 資料作成,Claude 業務効率化）")
    selection = input("> ").strip()

    extra_kws = []
    if not selection:
        return []

    # 番号選択か直接入力か判定
    if re.match(r'^[\d,\s]+$', selection):
        nums = [int(n.strip()) for n in selection.split(",") if n.strip().isdigit()]
        for n in nums:
            if 1 <= n <= len(candidates):
                word = candidates[n - 1][0]
                # メインKWと組み合わせる
                extra_kws.append(f"{main_kw} {word}")
    else:
        extra_kws = [k.strip() for k in selection.split(",") if k.strip()]

    return extra_kws[:4]


# ============================================================
# 市場マップ MD 生成
# ============================================================

def _fmt_num(n) -> str:
    """数値を3桁カンマ区切りに"""
    if isinstance(n, float):
        return f"{n:,.0f}"
    return f"{int(n):,}"


def generate_market_map(all_results: dict, keywords: list, output_dir: Path) -> Path:
    """検索結果から市場マップMDを生成する。KW別に再生数上位10本を出力。"""
    today = datetime.now().strftime("%Y-%m-%d")
    kw_label = "".join(f"「{kw}」" for kw in keywords)

    # 統合サマリ用に全動画を集約
    all_items = []
    for items in all_results.values():
        all_items.extend(items)

    lines = []
    lines.append("# リサーチ結果 — 市場マップ\n")
    lines.append(f"**生成日**: {today}")
    lines.append(f"**検索キーワード**: {kw_label}\n")
    lines.append("---\n")

    # 統合サマリ
    if all_items:
        views = sorted([i["viewCount"] for i in all_items])
        like_rates = sorted([i["likeRate"] * 100 for i in all_items])
        comment_rates = sorted([i["commentRate"] * 100 for i in all_items])
        vs_ratios = sorted([i["viewSubRatio"] for i in all_items])
        daily_avgs = sorted([i["dailyAvg"] for i in all_items])

        def median(lst):
            n = len(lst)
            if n == 0:
                return 0
            mid = n // 2
            return (lst[mid - 1] + lst[mid]) / 2 if n % 2 == 0 else lst[mid]

        lines.append("## 統合サマリ\n")
        lines.append("| 項目 | 値 |")
        lines.append("|:---|:---|")
        lines.append(f"| 総動画数 | {len(all_items)}本 |")
        lines.append(f"| 再生数中央値 | {_fmt_num(median(views))} |")
        lines.append(f"| 高評価率中央値 | {median(like_rates):.2f}% |")
        lines.append(f"| コメント率中央値 | {median(comment_rates):.3f}% |")
        lines.append(f"| View/Sub比中央値 | {median(vs_ratios):.2f} |")
        lines.append(f"| 日次平均再生中央値 | {_fmt_num(median(daily_avgs))} |")
        lines.append("\n---\n")

    # KW別 上位10本
    for kw, items in all_results.items():
        top10 = sorted(items, key=lambda x: x["viewCount"], reverse=True)[:10]
        lines.append(f"## KW「{kw}」（{len(items)}本中 上位10本）\n")

        for rank, item in enumerate(top10, 1):
            dur_sec = parse_duration(item["duration"])
            warn = " ⚠️" if 0 < dur_sec < 300 else ""
            lines.append(f"### #{rank} {item['title']}{warn}\n")
            if item.get("thumbnailUrl"):
                lines.append(f'<img src="{item["thumbnailUrl"]}" width="450">\n')
            lines.append("| 項目 | 値 |")
            lines.append("|:---|:---|")
            lines.append(f"| チャンネル名 | {item['channelTitle']} |")
            lines.append(f"| 公開日 | {item['publishedAt']} |")
            lines.append(f"| 登録者数 | {_fmt_num(item['subscriberCount'])} |")
            lines.append(f"| 再生数 | {_fmt_num(item['viewCount'])} |")
            lines.append(f"| いいね数 | {_fmt_num(item['likeCount'])} |")
            lines.append(f"| コメント数 | {_fmt_num(item['commentCount'])} |")
            lines.append(f"| 高評価率 | {item['likeRate'] * 100:.2f}% |")
            lines.append(f"| コメント率 | {item['commentRate'] * 100:.3f}% |")
            lines.append(f"| V/S比 | {item['viewSubRatio']:.2f} |")
            lines.append(f"| 日次平均再生 | {_fmt_num(item['dailyAvg'])} |")
            lines.append(f"| 尺 | {item['durationFmt']} |")
            lines.append(f"| 動画URL | {item['videoUrl']} |")
            lines.append("")

    md_path = output_dir / "market-map.md"
    md_path.write_text("\n".join(lines), encoding="utf-8")
    return md_path


# ============================================================
# メインフロー
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="YouTube リサーチ → Googleスプレッドシート出力")
    parser.add_argument("--keywords", "-k", help="カンマ区切りキーワード")
    parser.add_argument("--ep", help="EP番号（例: EP-05）")
    parser.add_argument("--period", type=int, default=6, help="検索期間（月数、デフォルト6）")
    parser.add_argument("--type", dest="video_type", choices=["any", "normal", "shorts"], default="normal")
    parser.add_argument("--no-comments", action="store_true")
    parser.add_argument("--market-map", action="store_true", help="市場マップMDを自動生成")
    parser.add_argument("--output-dir", default="./output", help="Excel・JSON出力先（デフォルト: ./output）")
    args = parser.parse_args()

    # --output-dir にJST実行時刻サフィックスを常に付与する
    # 形式: {base}-time{HHMM}（例: Claude-Code-使い方-Research-20260409-time1827）
    # LLMが -time サフィックスを既に付けている場合は二重付与しない
    # （LLMの判断に依存せず、スクリプト側で決定論的に一意化）
    JST = timezone(timedelta(hours=9))
    jst_suffix = datetime.now(JST).strftime("time%H%M")
    base_output_dir = Path(args.output_dir)
    if "-time" in base_output_dir.name:
        output_dir = base_output_dir
    else:
        output_dir = base_output_dir.parent / f"{base_output_dir.name}-{jst_suffix}"
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"📁 出力フォルダ: {output_dir}")

    # キーワード決定
    keywords = []
    keywords_from_cli = False  # --keywords で明示指定された場合はA/Bプロンプトをスキップ
    if args.keywords:
        keywords = [k.strip() for k in args.keywords.split(",") if k.strip()]
        keywords_from_cli = True
    elif args.ep:
        plan = find_plan_file(args.ep)
        if plan:
            print(f"企画骨子: {plan.name}")
            keywords = extract_keywords_from_plan(plan)
            if keywords:
                print(f"抽出KW ({len(keywords)}件): {', '.join(keywords)}")
            else:
                print("キーワードが見つかりませんでした。")

    if not keywords:
        raw = input("検索キーワードを入力（カンマ区切り）: ").strip()
        keywords = [k.strip() for k in raw.split(",") if k.strip()]

    if not keywords:
        print("キーワードがありません。終了します。")
        sys.exit(1)

    # 日付範囲（関連KW候補生成でも使うので先に計算）
    end_dt = datetime.now()
    start_dt = end_dt - timedelta(days=30 * args.period)
    start_str = start_dt.strftime("%Y-%m-%d")
    end_str = end_dt.strftime("%Y-%m-%d")
    date_label = datetime.now().strftime("%Y%m%d")

    # API初期化
    env = load_env(ENV_PATH)
    api_keys = collect_api_keys(env)
    yt = YouTubeClient(api_keys)

    # 1KWのみ → 関連KW自動候補生成（--keywords指定時はスキップ）
    if len(keywords) == 1 and not keywords_from_cli:
        print(f"\nメインKW: 「{keywords[0]}」")
        print("A) 単一KWでリサーチ（1本検索）")
        print("B) 関連KWも含めてリサーチ（計5本検索）— トレンドKW自動候補あり")
        choice = input("選択 [A/B]: ").strip().upper()
        if choice == "B":
            extra_kws = suggest_related_keywords(yt, keywords[0], start_str, end_str)
            if not extra_kws:
                print(f"\n関連KWを4本入力してください（カンマ区切り）:")
                extra = input("> ").strip()
                extra_kws = [k.strip() for k in extra.split(",") if k.strip()]
            keywords = keywords + extra_kws[:4]
            print(f"\n実行するKW ({len(keywords)}本):")
            for idx, kw in enumerate(keywords, 1):
                print(f"  {idx}. {kw}")
            print("\nA) このKWで実行")
            print("B) 変更する — 修正後のKWをカンマ区切りで記入して送信してください")
            confirm = input("> ").strip()
            if confirm.upper() == "A" or confirm == "":
                pass  # そのまま実行
            else:
                # A以外の入力はKW修正として扱う
                keywords = [k.strip() for k in confirm.split(",") if k.strip()]
                if not keywords:
                    sys.exit(0)
                print(f"\n実行するKW ({len(keywords)}本):")
                for idx, kw in enumerate(keywords, 1):
                    print(f"  {idx}. {kw}")

    print(f"\n検索期間: {start_str} 〜 {end_str} | タイプ: {args.video_type} | APIキー: {len(api_keys)}本\n")

    # 検索実行
    all_results = {}
    for i, kw in enumerate(keywords, 1):
        print(f"[{i}/{len(keywords)}] 「{kw}」を検索中...")
        try:
            search_items = yt.search(kw, start_str, end_str)
            if not search_items:
                print("  → 0件（スキップ）")
                continue
            video_ids = [s["id"]["videoId"] for s in search_items]
            channel_ids = [s["snippet"]["channelId"] for s in search_items]
            print(f"  メタデータ取得中... ({len(video_ids)}件)")
            vd = yt.get_video_details(video_ids)
            cd = yt.get_channel_details(channel_ids)
            items = build_items(search_items, vd, cd, args.video_type)
            all_results[kw] = items
            print(f"  → {len(items)}件（フィルター後）")
        except Exception as e:
            print(f"  エラー: {e}")

    if not all_results:
        print("有効な検索結果がありません。終了します。")
        sys.exit(1)

    print("\n" + "=" * 50)
    print("■ 検索結果サマリ")
    for kw, items in all_results.items():
        print(f"  「{kw}」: {len(items)}件")
    print("=" * 50 + "\n")

    # コメント取得（KWごとに分離して管理）
    comments_by_kw = {}
    if not args.no_comments:
        if keywords_from_cli:
            do_comments = True
        else:
            try:
                print("A) コメントも取得する")
                print("B) コメントは取得しない")
                ans = input("選択 [A/B]: ").strip().upper()
                do_comments = ans != "B"
            except EOFError:
                do_comments = True
        if do_comments:
            for kw, items in all_results.items():
                # VS比≧1.0 AND 日次平均≧100 AND 尺≧300秒 の上位10本
                candidates = [
                    item for item in items
                    if item["viewSubRatio"] >= 1.0
                    and item["dailyAvg"] >= 100
                    and parse_duration(item["duration"]) >= 300
                ]
                candidates.sort(key=lambda x: x["viewCount"], reverse=True)
                top = candidates[:10]
                if not top:
                    continue
                kw_comments = []
                print(f"\n「{kw}」コメント取得対象: {len(top)}本")
                for j, v in enumerate(top, 1):
                    print(f"  [{j}/{len(top)}] {v['title'][:50]}...")
                    threads = yt.get_comments(v["videoId"])
                    for t in threads:
                        kw_comments.append({
                            "videoId": v["videoId"],
                            "videoTitle": v["title"],
                            "videoUrl": v["videoUrl"],
                            "author": t["author"],
                            "text": t["text"],
                            "likes": t["likes"],
                            "publishedAt": t["publishedAt"],
                            "replies": t["replies"],
                        })
                    print(f"     → {len(threads)}スレッド取得")
                if kw_comments:
                    comments_by_kw[kw] = kw_comments

    # ============================================================
    # Excel 出力（KWごとに個別ファイル）
    # ============================================================
    print("\nExcelファイルを作成中...")
    created_files = []  # (kw, filename, filepath) のリスト

    for idx, (kw, items) in enumerate(all_results.items(), 1):
        file_name = generate_spreadsheet_name(kw, idx, date_label) + ".xlsx"
        file_path = output_dir / file_name
        print(f"  [{idx}/{len(all_results)}] {file_name} ({len(items)}件)")

        wb = Workbook()
        ws = wb.active
        ws.title = "リサーチデータ"
        write_research_sheet(ws, items, kw)

        if kw in comments_by_kw:
            ws_cmt = wb.create_sheet("コメント")
            write_comments_sheet(ws_cmt, comments_by_kw[kw])
            print(f"    + コメントシート: {len(comments_by_kw[kw])}件")

        wb.save(file_path)
        created_files.append((kw, file_name, file_path))

    # ============================================================
    # コメントJSON出力（KWごとに個別ファイル）
    # ============================================================
    json_files = []
    for idx, (kw, _items) in enumerate(all_results.items(), 1):
        if kw not in comments_by_kw:
            continue
        json_name = f"{generate_spreadsheet_name(kw, idx, date_label)}_comments.json"
        json_path = output_dir / json_name
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({
                "keyword": kw,
                "date": date_label,
                "total_threads": len(comments_by_kw[kw]),
                "comments": comments_by_kw[kw],
            }, f, ensure_ascii=False, indent=2)
        json_files.append((kw, json_name, json_path))
        print(f"  JSON出力: {json_name}")

    # ============================================================
    # 完了サマリ
    # ============================================================
    print("\n" + "=" * 60)
    print("■ 完了")
    print("=" * 60)
    print(f"\nExcelファイル ({len(created_files)}件):")
    for kw, name, path in created_files:
        print(f"  {name}")
        print(f"    KW: {kw}")
        print(f"    Path: {path}")
    if json_files:
        print(f"\nコメントJSON ({len(json_files)}件):")
        for kw, name, path in json_files:
            print(f"  {name}")
            print(f"    → {path}")

    # ============================================================
    # 市場マップ MD 生成
    # ============================================================
    do_market_map = False
    if args.market_map:
        do_market_map = True
    elif keywords_from_cli:
        pass  # --keywords指定 + --market-map未指定 → 生成しない
    else:
        try:
            print("\nA) リサーチ結果をMarkdownで一覧表示する（market-map.md）")
            print("B) 不要")
            ans = input("選択 [A/B]: ").strip().upper()
            do_market_map = ans != "B"
        except EOFError:
            pass

    if do_market_map:
        print("\n市場マップMDを生成中...")
        md_path = generate_market_map(all_results, keywords, output_dir)
        print(f"  → {md_path}")

    print()


if __name__ == "__main__":
    main()
